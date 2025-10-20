import csv
import tkinter as tk
from tkinter import font
import json
import os
import matplotlib.pyplot as plt
from collections import defaultdict
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from tkinter import filedialog
import psycopg2
from psycopg2.extras import RealDictCursor

load_dotenv()

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "database": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASS")
}

FILENAME = "expenses.json"

total_label = None
message_label = None
root = None
search_entry = None
listbox = None
category_entry = None
amount_entry = None
CURRENT_YEAR = 2025
USE_DATABASE = True


def connect_db():
    try:
        return psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        print(f"DB connection error: {e}")
        return None

def db_add_expense(category, amount):
    conn = connect_db()
    if not conn: return
    cur = conn.cursor()
    cur.execute("INSERT INTO expenses (category, amount, date) VALUES (%s, %s, CURRENT_DATE)", (category, amount))
    conn.commit()
    cur.close()
    conn.close()

def db_get_all_expenses():
    conn = connect_db()
    if not conn: return []
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, category, amount, date FROM expenses ORDER BY id ASC")
    data = cur.fetchall()
    cur.close()
    conn.close()
    return data

def db_delete_expense(index):
    data = db_get_all_expenses()
    if not data or index >= len(data): return False
    target_id = data[index]['id']
    conn = connect_db()
    if not conn: return False
    cur = conn.cursor()
    cur.execute("DELETE FROM expenses WHERE id = %s", (target_id,))
    conn.commit()
    cur.close()
    conn.close()
    return True

def load_json():
    if not os.path.exists(FILENAME): return []
    try:
        with open(FILENAME, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError:
        return []


def save_json(expenses):
    with open(FILENAME, "w", encoding="utf-8") as f:
        json.dump(expenses, f, indent=4, ensure_ascii=False)

def load_data():
    db_data = db_get_all_expenses()
    if db_data:
        return [{"category": x["category"], "amount": x["amount"], "date": str(x["date"])} for x in db_data]
    else:
        return load_json()

def sync_to_json():
    db_data = db_get_all_expenses()
    if db_data:
        clean_data = [{"category": x["category"], "amount": x["amount"], "date": str(x["date"])} for x in db_data]
        save_json(clean_data)
        print("JSON synced from PostgreSQL")
    else:
        print("No DB data found to sync")


def sync_to_db():
    db_data = db_get_all_expenses()
    if db_data:
        print("DB already has data, skipping JSON→DB sync.")
        return
    json_data = load_json()
    if not json_data:
        print("No JSON data to sync into DB.")
        return

    conn = connect_db()
    if not conn: return
    cur = conn.cursor()
    for item in json_data:
        cur.execute(
            "INSERT INTO expenses (category, amount, date) VALUES (%s, %s, %s)",
            (item["category"], item["amount"], item.get("date", datetime.now().date()))
        )
    conn.commit()
    cur.close()
    conn.close()
    print("PostgreSQL synced from JSON")

def save_data(expenses):
    with open(FILENAME, "w", encoding="utf-8") as f:
        json.dump(expenses, f, indent=4, ensure_ascii=False)

def add_expense(category, amount):
    db_add_expense(category, amount)
    sync_to_json()


def delete_expense(index):
    success = db_delete_expense(index)
    if success:
        sync_to_json()

def calculate_total():
    expenses = load_data()
    total = sum(item['amount'] for item in expenses)
    return total


def show_flash_message(message, color="black"):
    global root
    if message_label and root:
        bg_color = 'light green' if color == 'green' else ('#FFDDE0' if color == 'red' else 'light blue')
        fg_color = 'green' if color == 'green' else ('red' if color == 'red' else 'blue')

        message_label.config(text=message, fg=fg_color, bg=bg_color, relief=tk.RIDGE, bd=2, padx=5, pady=2)
        root.after(3000, lambda: message_label.config(text="", fg="black", bg=root['bg'], relief=tk.FLAT, bd=0))


def show_total():
    total = calculate_total()
    if total_label:
        total_label.config(text=f" Total: {total} Tenge")


def update_listbox():
    listbox.delete(0, tk.END)
    expenses = load_data()
    for item in expenses:
        date_str = item.get("date", "N/A")
        listbox.insert(tk.END, f"{item['category']} - {item['amount']} Tenge | {date_str}")

    if search_entry:
        search_entry.delete(0, tk.END)

def show_stats():
    expenses = load_data()

    if not expenses:
        show_flash_message("No data found.", "red")
        return

    max_cat = max(expenses, key=lambda x: x['amount'])
    min_cat = min(expenses, key=lambda x: x['amount'])

    msg = f" Max: {max_cat['category']} ({max_cat['amount']} Tenge)  |   Min: {min_cat['category']} ({min_cat['amount']} Tenge)"
    show_flash_message(msg, "blue")


def on_add_expense():
    category = category_entry.get().strip()
    amount_str = amount_entry.get().strip()

    if not category or not amount_str:
        show_flash_message("Error: Fill in all fields!", "red")
        return

    try:
        amount = int(amount_str)
        if amount <= 0:
            show_flash_message("Error: Amount must be a positive number!", "red")
            return
    except ValueError:
        show_flash_message("Error: Amount must be an integer!", "red")
        return

    add_expense(category, amount)

    category_entry.delete(0, tk.END)
    amount_entry.delete(0, tk.END)
    update_listbox()
    show_total()
    show_flash_message(f"'{category}' added!", "green")


def on_delete_expense():
    selected = listbox.curselection()
    if not selected:
        show_flash_message("Error: Select an item to delete!", "red")
        return

    index = selected[0]

    deleted_item = delete_expense(index)

    if deleted_item:
        update_listbox()
        show_total()
        show_flash_message(f"Deleted: {deleted_item['category']} - {deleted_item['amount']} Tenge", "blue")


def search_expense():
    query = search_entry.get().strip().lower()

    listbox.delete(0, tk.END)
    expenses = load_data()
    found = False

    if not query:
        update_listbox()
        return

    for item in expenses:
        exp_str = f"{item['category']} - {item['amount']} Tenge | {item.get('date', 'N/A')}"
        if query in exp_str.lower():
            listbox.insert(tk.END, exp_str)
            found = True

    if not found:
        listbox.insert(tk.END, "Nothing found.")
        show_flash_message("Nothing found.", "orange")

def show_pie_chart():
    expenses = load_data()

    if not expenses:
        show_flash_message("No data to show chart.", "red")
        return

    category_totals = defaultdict(int)
    for item in expenses:
        category_totals[item['category']] += item['amount']

    labels = category_totals.keys()
    sizes = category_totals.values()

    total_amount = sum(sizes)
    threshold = total_amount * 0.05

    other_size = 0
    final_labels = []
    final_sizes = []

    for label, size in zip(labels, sizes):
        if size < threshold and len(labels) > 8:
            other_size += size
        else:
            final_labels.append(label)
            final_sizes.append(size)

    if other_size > 0:
        final_labels.append("Other")
        final_sizes.append(other_size)

    fig1, ax1 = plt.subplots()
    ax1.pie(final_sizes, labels=final_labels, autopct='%1.1f%%', startangle=90,
            textprops={'fontsize': 8}, wedgeprops={'edgecolor': 'black', 'linewidth': 1, 'antialiased': True})

    ax1.axis('equal')

    plt.title(f"Expense Distribution by Category - {CURRENT_YEAR}", pad=20)
    plt.show()

def export_to_csv():
    expenses = load_data()
    if not expenses:
        show_flash_message("No data to export!", "red")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV files", "*.csv")],
        title="Save CSV File"
    )

    if not file_path:
        return

    try:
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = ["category", "amount", "date"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(expenses)

        show_flash_message(f"Data exported to: {os.path.basename(file_path)}", "green")
    except Exception as e:
        show_flash_message(f"Export failed: {e}", "red")

def import_from_csv():
    file_path = filedialog.askopenfilename(
        filetypes=[("CSV files", "*.csv")],
        title="Select CSV File to Import"
    )

    if not file_path:
        return

    try:
        with open(file_path, "r", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            imported_expenses = []
            for row in reader:
                try:
                    amount = int(row["amount"])
                    category = row["category"].strip()
                    date = row.get("date", "N/A")
                    imported_expenses.append({
                        "category": category,
                        "amount": amount,
                        "date": date
                    })
                except (ValueError, KeyError):
                    continue

        if imported_expenses:
            existing_data = load_data()
            existing_data.extend(imported_expenses)
            save_data(existing_data)
            update_listbox()
            show_total()
            show_flash_message(f"Imported {len(imported_expenses)} records from CSV!", "green")
        else:
            show_flash_message("No valid records found in CSV!", "orange")

    except Exception as e:
        show_flash_message(f"Import failed: {e}", "red")

def export_to_excel():
    expenses = load_data()
    if not expenses:
        show_flash_message("No data to export!", "red")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Excel File"
    )

    if not file_path:
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        ws.append(["Category", "Amount", "Date"])

        for item in expenses:
            ws.append([item["category"], item["amount"], item.get("date", "N/A")])

        for column in ws.columns:
            max_length = 0
            col = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col].width = adjusted_width

        wb.save(file_path)
        show_flash_message(f"Data exported to: {os.path.basename(file_path)}", "green")

    except Exception as e:
        show_flash_message(f"Export failed: {e}", "red")

def import_from_excel():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")],
        title="Select Excel File to Import"
    )

    if not file_path:
        return

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        imported_expenses = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            category, amount, date = row
            try:
                amount = int(amount)
                imported_expenses.append({
                    "category": str(category).strip(),
                    "amount": amount,
                    "date": str(date) if date else "N/A"
                })
            except (ValueError, TypeError):
                continue

        if imported_expenses:
            existing_data = load_data()
            existing_data.extend(imported_expenses)
            save_data(existing_data)
            update_listbox()
            show_total()
            show_flash_message(f"Imported {len(imported_expenses)} records from Excel!", "green")
        else:
            show_flash_message("No valid records found in Excel file.", "orange")

    except Exception as e:
        show_flash_message(f"Import failed: {e}", "red")


def setup_ui():
    global root, total_label, message_label, category_entry, amount_entry, listbox, search_entry

    root = tk.Tk()
    root.title("Expense Tracker")
    root.geometry("480x655")
    root.resizable(False, False)
    root.configure(bg='#f0f0f0')

    button_font = font.Font(family="Arial", size=10, weight="bold")

    input_frame = tk.LabelFrame(root, text="Add/Delete", padx=10, pady=10, bg='#ffffff', fg='#333333')
    input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

    tk.Label(input_frame, text="Category:", bg='#ffffff').grid(row=0, column=0, padx=5, pady=5, sticky="w")
    category_entry = tk.Entry(input_frame, width=30)
    category_entry.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(input_frame, text="Amount (Tenge):", bg='#ffffff').grid(row=1, column=0, padx=5, pady=5, sticky="w")
    amount_entry = tk.Entry(input_frame, width=30)
    amount_entry.grid(row=1, column=1, padx=5, pady=5)

    button_frame = tk.Frame(input_frame, bg='#ffffff')
    button_frame.grid(row=2, column=0, columnspan=2, pady=10)

    top_row = tk.Frame(button_frame, bg='#ffffff')
    top_row.pack(pady=3)

    tk.Button(top_row, text="Add", command=on_add_expense, bg='#4CAF50', fg='white', width=10,
              font=button_font).pack(side=tk.LEFT, padx=5)
    tk.Button(top_row, text="Delete", command=on_delete_expense, bg='#f44336', fg='white', width=10,
              font=button_font).pack(side=tk.LEFT, padx=5)
    tk.Button(top_row, text="Stats", command=show_stats, bg='#2196F3', fg='white', width=10,
              font=button_font).pack(side=tk.LEFT, padx=5)
    tk.Button(top_row, text="Chart (π)", command=show_pie_chart, bg='#FFC107', fg='black', width=10,
              font=button_font).pack(side=tk.LEFT, padx=5)

    bottom_row = tk.Frame(button_frame, bg='#ffffff')
    bottom_row.pack(pady=3)

    tk.Button(bottom_row, text="Export CSV", command=export_to_csv, bg='#9C27B0', fg='white', width=10,
              font=button_font).pack(side=tk.LEFT, padx=5)
    tk.Button(bottom_row, text="Import CSV", command=import_from_csv, bg='#009688', fg='white', width=10,
              font=button_font).pack(side=tk.LEFT, padx=5)
    tk.Button(bottom_row, text="Export XLSX", command=export_to_excel, bg='#673AB7', fg='white', width=12,
              font=button_font).pack(side=tk.LEFT, padx=5)
    tk.Button(bottom_row, text="Import XLSX", command=import_from_excel, bg='#3F51B5', fg='white', width=12,
              font=button_font).pack(side=tk.LEFT, padx=5)

    search_frame = tk.LabelFrame(root, text="Search", padx=10, pady=5, bg='#f0f0f0', fg='#333333')
    search_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

    search_inner = tk.Frame(search_frame, bg='#f0f0f0')
    search_inner.pack(anchor="center", pady=5)

    search_entry = tk.Entry(search_inner, width=35, justify="center")
    search_entry.pack(side=tk.LEFT, padx=5)

    tk.Button(search_inner, text="Search", command=search_expense, bg="#FF9800", fg="white", width=10,
              font=button_font).pack(side=tk.LEFT, padx=5)

    list_container_frame = tk.LabelFrame(root, text="Expense List", padx=5, pady=5, bg='#ffffff', fg='#333333')
    list_container_frame.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

    listbox = tk.Listbox(list_container_frame, width=50, height=12, bg='#f9f9f9', fg='#000000',
                         selectbackground='#b3e0ff')
    listbox.pack(side="left", fill="y", expand=True)

    scrollbar = tk.Scrollbar(list_container_frame, orient="vertical", command=listbox.yview)
    scrollbar.pack(side="right", fill="y")
    listbox.config(yscrollcommand=scrollbar.set)

    message_label = tk.Label(root, text="", font=("Arial", 10), wraplength=400, bg=root['bg'])
    message_label.grid(row=3, column=0, pady=5)

    total_label = tk.Label(root, text="Total: 0 Tenge", font=("Arial", 14, "bold"), fg='#006400', bg=root['bg'])
    total_label.grid(row=4, column=0, pady=10)

    footer_label = tk.Label(root, text=f"© {CURRENT_YEAR} Zhassar Otan | Data: {FILENAME}",
                            font=("Arial", 8), fg='#666666', bg=root['bg'])
    footer_label.grid(row=5, column=0, pady=5)

    update_listbox()
    show_total()

    root.mainloop()

if __name__ == "__main__":
    setup_ui()